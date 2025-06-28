from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
from datetime import datetime
from selenium.webdriver.chrome.service import Service

import time

wb = load_workbook('data.xlsx')
sheet = wb.active
service=Service(r"D:\chromedriver-win64\chromedriver-win64\chromedriver.exe")
driver = webdriver.Chrome(service=service)
driver.maximize_window()
wait = WebDriverWait(driver, 15)

for i in range(2, sheet.max_row + 1):
    origin = sheet.cell(row=i, column=1).value
    destination = sheet.cell(row=i, column=2).value
    travel_date = sheet.cell(row=i, column=3).value
    passengers = sheet.cell(row=i, column=4).value
    airline = sheet.cell(row=i, column=5).value

    driver.get("https://www.cleartrip.com/")
    wait.until(EC.presence_of_element_located((By.XPATH, "//input[@placeholder='Where from?']")))

    from_input = driver.find_element(By.XPATH, "//input[@placeholder='Where from?']")
    from_input.click()
    ActionChains(driver).send_keys(origin).pause(1).send_keys("\n").perform()

    to_input = driver.find_element(By.XPATH, "//input[@placeholder='Where to?']")
    to_input.click()
    ActionChains(driver).send_keys(destination).pause(1).send_keys("\n").perform()

    driver.find_element(By.XPATH, "//div[contains(text(), 'Departure')]").click()
    time.sleep(1)
    try:
        date_str = travel_date.strftime("%Y-%m-%d")
        date_button = driver.find_element(By.XPATH, f"//div[@aria-label='{date_str}']")
        date_button.click()
    except:
        print(f"Date not clickable for: {travel_date}")
        continue

    driver.find_element(By.XPATH, "//div[contains(text(),'Travellers & Class')]").click()
    time.sleep(1)
    for _ in range(passengers - 1):
        driver.find_element(By.XPATH, "//button[contains(@aria-label,'add adult')]").click()
        time.sleep(0.5)
    driver.find_element(By.XPATH, "//button[text()='Done']").click()

    driver.find_element(By.XPATH, "//button[text()='Search flights']").click()
    time.sleep(10)

    print(f"Searched: {origin} to {destination} on {travel_date}, {passengers} passengers, Airline: {airline}")

driver.quit()
