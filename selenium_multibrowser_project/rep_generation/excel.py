from openpyxl import Workbook, load_workbook
import os

def write_to_excel(log_data):
    filepath = r"D:\Desktop\selenium_testing\selenium_multibrowser_project\docs\Test_Cases.xlsx"
    folder_path = os.path.dirname(filepath)
    if not os.path.exists(folder_path):
        os.makedirs(folder_path)
    if os.path.exists(filepath):
        wb = load_workbook(filepath)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Automation Logs"
        ws.append(["Timestamp", "Module", "Step Description", "Status", "Details"])
    for row in log_data:
        ws.append(row)
    wb.save(filepath)
